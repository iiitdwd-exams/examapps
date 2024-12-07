import pandas as pd
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection, PatternFill
from openpyxl.formatting.rule import CellIsRule

df = pd.read_csv("FList_2024_05.csv")
g = df.groupby("roll_no")
chunks = [group for _, group in g]
x = chunks[1].copy()
x["register"] = 0
r = x.iloc[0, 0].lower()
n = x.iloc[0, 1].lower().replace(" ", "_")
fname = f"{r}_{n}.xlsx"
x.loc[:, ["roll_no", "name", "code", "course", "credits", "register"]].to_excel(
    fname, index=False
)

wb = xl.load_workbook(fname)
ws = wb.active
if ws:
    last_row = ws.max_row

    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    formula_cell = ws.cell(row=last_row + 1, column=5)
    formula_cell.value = f'=SUMIF(F2:F{last_row}, ">0", E2:E{last_row})'
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    rule_green = CellIsRule(operator="lessThanOrEqual", formula=["12"], fill=green_fill)
    rule_yellow = CellIsRule(operator="equal", formula=["13"], fill=green_fill)
    rule_red = CellIsRule(operator="greaterThan", formula=["13"], fill=green_fill)
    ws.conditional_formatting.add(f"E{last_row+1}", rule_green)
    ws.conditional_formatting.add(f"E{last_row+1}", rule_yellow)
    ws.conditional_formatting.add(f"E{last_row+1}", rule_red)
    for row in ws.iter_rows(min_col=6, max_col=6):
        for cell in row:
            cell.protection = Protection(locked=False)
    ws.cell(row=last_row + 1, column=5).protection = Protection(locked=False)
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                width = len(str(cell.value))
                if width > max_length:
                    max_length = width
            except Exception as e:
                print(f"Exception: {e}")
        print(f"{column}: {max_length}")
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    ws.protection.set_password("secret")
wb.save(fname)
